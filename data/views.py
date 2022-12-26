from rest_framework.decorators import api_view
from rest_framework.response import Response
from . models import *
from user.models import *
from datetime import datetime, timedelta, date
import pytz
from django.shortcuts import render, redirect, HttpResponse
from django.contrib.auth.decorators import login_required
from .serializer import DataSerializer
from . import documentor
import io
from django.contrib import messages
import openpyxl
from django.http import HttpResponse
import os


@api_view(['GET'])
def get_strain_data(request, value):
    data_v = request.data
    dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
    value_float = float(value)

    v_status = False
    v_cond = False

    if value_float >= 50:
        pass
    elif value_float >= 40 and value_float < 50:
        v_cond = True
    else:
        v_status = True

    data_value = Data.objects.create(
        strain = value_float,
        created = dartime,
        condition = v_cond,
        status = v_status
    )

    serializer = DataSerializer(data_value, many=False)
    print(data_value)
    return Response(serializer.data)


@login_required(login_url='user:login')
def bridge_info(request):
    if request.method == 'POST':
        name = request.POST.get('name_bridge')
        bridge_type = request.POST.get('bridge_type')
        no_of_lanes = request.POST.get('no_of_lanes')
        lane_width = request.POST.get('laneWidth')
        pedestrian_lane = request.POST.get('pedestrianlane')
        effective_span = request.POST.get('effectivespan')
        max_sema_dif = request.POST.get('max_sema_dif')
        location = request.POST.get('location')

        print(name, bridge_type, no_of_lanes, lane_width, pedestrian_lane, effective_span, max_sema_dif, location)
    return render(request, 'data/bridge_information.html', {})


@login_required(login_url='user:login')
def one_bridge_info(request, val):
    if val:
        value = int(val)
        if value == 1:
            name = "XYZ Bridge"
            q_fetch = Bridge.objects.get(name=name)
        
        if request.method == 'POST':
            name = "XYZ Bridge"
            q_fetch = Bridge.objects.get(name=name)
            # name_changed = request.POST.get('name_bridge')
            q_fetch.bridge_type = request.POST.get('bridge_type')
            q_fetch.lanes_number = request.POST.get('no_of_lanes')
            q_fetch.lane_width = request.POST.get('laneWidth')
            q_fetch.ped_lane_width = request.POST.get('pedestrianlane')
            q_fetch.effective_span = request.POST.get('effectivespan')
            q_fetch.max_sema_deflection = request.POST.get('max_sema_dif')
            q_fetch.location = request.POST.get('location')
            q_fetch.save()
            
        return render(request, 'data/one_bridge_info.html', {'q_data': q_fetch})
    else:
        return redirect('data:bridge_info' )
    


@login_required(login_url='user:login')
def bridge_inve(request):
    if request.method == "POST" and request.FILES['upload']:
        _complete, _incomplete, _count, excel_file = False, False, 0, request.FILES['upload']
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["example"]
        print(worksheet)
        # iterating over the rows and
        # getting value from each cell in row
        excel_data = list()                     
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            
            if row_data[0] != 'None' or row_data[1] != 'MAINTENANCE HISTORY':
                if row_data[0] != 'S/N' and row_data[1] != 'DATE' and row_data[2] != 'DEFECTS  NOTED':
                    if row_data[1] != 'None' and row_data[3] != 'None' and row_data[4] != 'None' and row_data[5] != 'None':
                        if row_data[0] != 'None' and row_data[2] != 'None':
                            main_activity = row_data[2]
                        
                        date = str(row_data[1])[:-8]
                        tasks = row_data[3]
                        amount = row_data[4]
                        status = row_data[5]
                    
                        InventoryData.objects.create(
                            date = date,
                            main_activity = main_activity,
                            activities = tasks,
                            cost = amount,
                            status = status
                        )
                        _complete = True
                        _count += 1
            
            else:
                _incomplete = True
        
        if _complete:
            messages.info(request, f'{_count} rows of data was Inserted !')
        
        if _incomplete:
            messages.info(request, f'Choose the correct sheet, Or Insert the data properly')

        value_list = InventoryData.objects.values_list('main_activity', flat=True).distinct().order_by('-created')[:20]

        group_by_value = {}
        for value in value_list:
            group_by_value[value] = InventoryData.objects.filter(main_activity=value).values('main_activity',
             'activities', 'status', 'cost', 'date')

        compareror = "none"   
        return render(request, 'data/bridge_inventory.html', {'inventory_group': group_by_value, 'comp': compareror})
    else: 
        value_list = InventoryData.objects.values_list('main_activity', flat=True).distinct().order_by('-created')[:20]

        group_by_value = {}
        for value in value_list:
            group_by_value[value] = InventoryData.objects.filter(main_activity=value).values('main_activity',
             'activities', 'status', 'cost', 'date')
        
        compareror = "none"   
        return render(request, 'data/bridge_inventory.html', {'inventory_group': group_by_value, 'comp': compareror})


@login_required(login_url='user:login')
def export_write_xls(request):
    dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
    time1 = dartime.ctime()
    user = request.user
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="InventoryTemplate_on_{time1}_by_{user}.xlsx"'
    
    path = os.path.dirname("media\docs\ ")
    file = os.path.join(path, 'EbmsInventory.xlsx')
    print(f'hii apa path.....{path}...na {file}')
    rb = openpyxl.load_workbook(file)
    
    rb.save(response)
    return response


@login_required(login_url='user:login')
def deck_defl(request):
    if request.method == 'POST' and (request.POST.get('interval')) is not None:
        selected = int(request.POST.get('interval'))
        if selected == 1:
            dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
            today = dartime.date()
            day_datas = Data.objects.filter(created__date=today)
            print(f'...{today}.!')
            checker = 'day'
            heading = f"Today's"
            return render(request, 'data/deck_deflection.html', {'disp_items':day_datas, 'interval':checker, 'head':heading})
        elif selected == 2:
            dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
            # now_time = datetime.now(tz=dartime)
            week_datas = DayData.objects.filter(created__gte=dartime-timedelta(days=7))
            checker = 'week'
            heading = "Last Seven Days'"
            print(f'It is a weeks data request ......{week_datas}')
            return render(request, 'data/deck_deflection.html', {'disp_items':week_datas, 'interval':checker, 'head':heading})
        elif selected == 3:
            dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
            month_datas = MonthData.objects.all()
            checker = 'month'
            heading = "Months"
            print(f'It is a month data request......{month_datas}')
            return render(request, 'data/deck_deflection.html', {'disp_items':month_datas, 'interval':checker, 'head':heading})
        elif selected == 4:
            dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
            year_datas = YearData.objects.all()
            checker = 'year'
            heading = "Year"
            print(f'It is a year data request......{year_datas}')
            return render(request, 'data/deck_deflection.html', {'disp_items':year_datas, 'interval':checker, 'head':heading})
        else:
            today = date.today()
            day_datas = Data.objects.filter(created__gte=today)
            return render(request, 'data/deck_deflection.html', {'disp_items':day_datas, 'interval':checker})
    else:
        return render(request, 'data/deck_deflection.html', {})


@login_required(login_url='user:login')
def daily_rep(request):
    #Getting data needed after submit the form in the page
    dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
    time1 = dartime.ctime()  # get curent time
    if int(dartime.day) == 1:
        if int(dartime.month) == 1:
            time2 = f'Last Day of Last Year  ( 00:00 - 23:59 )' # Tabulation Time
        else:
            time2 = f'{dartime.year}-{dartime.month-1}- Previous Day  ( 00:00 - 23:59 )' # Tabulation Time
    else:
        time2 = f'{dartime.year}-{dartime.month}-{dartime.day-1}  ( 00:00 - 23:59 )' # Tabulation Time

    checker, name = 'Daily', "XYZ Bridge"
    q_fetch = Bridge.objects.get(name=name)

    #Report creation
    word_doc = documentor.word_creator(checker, time1, time2, q_fetch) 
    doc_io = io.BytesIO()
    word_doc.save(doc_io)  # save to memory stream
    doc_io.seek(0)

    # rewind the stream
    response = HttpResponse(
    doc_io.getvalue(),
    content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )

    response["Content-Disposition"] = f'attachment; filename = "Daily_Report_({time1}).docx"'
    response["Content-Encoding"] = "UTF-8"
    return response


@login_required(login_url='user:login')
def weekly_rep(request):
    #Getting data needed after submit the form in the page
    dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
    time1 = dartime.ctime()  # get curent time
    
    time2 = f"Last Seven Days' Average in every ( 00:00 - 23:59 )" # Tabulation Time

    checker, name = 'Weekly', "XYZ Bridge"
    q_fetch = Bridge.objects.get(name=name)

    #Report creation
    word_doc = documentor.word_creator(checker, time1, time2, q_fetch) 
    doc_io = io.BytesIO()
    word_doc.save(doc_io)  # save to memory stream
    doc_io.seek(0)

    # rewind the stream
    response = HttpResponse(
    doc_io.getvalue(),
    content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )

    response["Content-Disposition"] = f'attachment; filename = "Weekly_Report_{time1}.docx"'
    response["Content-Encoding"] = "UTF-8"
    return response


@login_required(login_url='user:login')
def monthly_rep(request):
    #Getting data needed after submit the form in the page
    dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
    time1 = dartime.ctime()  # get curent time
    
    if int(dartime.month) == 1:
        time2 = f'{dartime.year - 1} - December   ( Ist - 31st December)' # Tabulation Time
    else:
        time2 = f'{dartime.year} - {dartime.month-1}   ( Average Data Monthly )' # Tabulation Time

    checker, name = 'Monthly', "XYZ Bridge"
    q_fetch = Bridge.objects.get(name=name)

    #Report creation
    word_doc = documentor.word_creator(checker, time1, time2, q_fetch) 
    doc_io = io.BytesIO()
    word_doc.save(doc_io)  # save to memory stream
    doc_io.seek(0)

    # rewind the stream
    response = HttpResponse(
    doc_io.getvalue(),
    content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )

    response["Content-Disposition"] = f'attachment; filename = "Monthly_Report_{time1}.docx"'
    response["Content-Encoding"] = "UTF-8"
    return response


@login_required(login_url='user:login')
def yearly_rep(request):
    #Getting data needed after submit the form in the page
    dartime = datetime.now(pytz.timezone('Africa/Dar_es_Salaam'))
    time1 = dartime.ctime()  #get curent time
        
    time2 = f'{dartime.year-1}   ( Prevoius Year )'  #Tabulation Time

    checker, name = 'Yearly', "XYZ Bridge"
    q_fetch = Bridge.objects.get(name=name)

    #Report creation
    word_doc = documentor.word_creator(checker, time1, time2, q_fetch) 
    doc_io = io.BytesIO()
    word_doc.save(doc_io)  # save to memory stream
    doc_io.seek(0)

    # rewind the stream
    response = HttpResponse(
    doc_io.getvalue(),
    content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )

    response["Content-Disposition"] = f'attachment; filename = "Yearly_Report_{time1}.docx"'
    response["Content-Encoding"] = "UTF-8"
    return response


@login_required(login_url='user:login')
def strain_page(request):
    strain_value = "kituuuu"
    print("inside strain page")
    return render(request, 'data/strain.html', {'strain': strain_value})


@login_required(login_url='user:login')
def index_page(request):
    first_bridge = Bridge.objects.all()[0]
    print(first_bridge.id)
    return render(request, 'data/index.html', {'first_bridge': first_bridge})


@login_required(login_url='user:login')
def displacement_view(request):
    context = {}
    return render(request, 'data/temp.html', context)


@login_required(login_url='user:login')
def err(request):
    return render(request, 'data/404.html', {})